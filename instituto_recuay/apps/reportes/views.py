from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, landscape, A4
from reportlab.lib import colors
from reportlab.lib.units import inch, cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import io
import qrcode
from PIL import Image as PILImage


@login_required
def reporte_asistencias_excel(request):
    if request.user.rol not in ['super_admin', 'admin', 'director', 'docente']:
        messages.error(request, 'Sin permisos para generar reportes.')
        return redirect('dashboard:admin_dashboard')

    from apps.asistencia.models import RegistroAsistencia
    from apps.academica.models import Curso

    registros = RegistroAsistencia.objects.select_related('alumno', 'curso').all()

    curso_id = request.GET.get('curso')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    if curso_id:
        registros = registros.filter(curso_id=curso_id)
    if fecha_inicio:
        registros = registros.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        registros = registros.filter(fecha__lte=fecha_fin)

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Asistencias"

    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='1B4F72', end_color='1B4F72', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = 'INSTITUTO SUPERIOR TECNOLOGICO PUBLICO REC'
    title_cell.font = Font(name='Calibri', bold=True, size=16, color='1B4F72')
    title_cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:H2')
    subtitle = ws['A2']
    subtitle.value = f'Reporte de Asistencias - Generado: {timezone.now().strftime("%d/%m/%Y %H:%M")}'
    subtitle.font = Font(name='Calibri', size=10, italic=True)
    subtitle.alignment = Alignment(horizontal='center')

    headers = ['N°', 'DNI', 'Alumno', 'Curso', 'Fecha', 'Hora Entrada', 'Estado', 'Metodo']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    estado_colores = {
        'P': '27AE60',
        'T': 'F39C12',
        'F': 'E74C3C',
        'J': '3498DB',
        'X': '8E44AD',
    }

    for idx, reg in enumerate(registros, 1):
        row = idx + 4
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=reg.alumno.dni).border = thin_border
        ws.cell(row=row, column=3, value=reg.alumno.get_full_name()).border = thin_border
        ws.cell(row=row, column=4, value=reg.curso.nombre).border = thin_border
        ws.cell(row=row, column=5, value=reg.fecha.strftime('%d/%m/%Y') if reg.fecha else '').border = thin_border
        ws.cell(row=row, column=6, value=str(reg.hora_entrada) if reg.hora_entrada else '').border = thin_border
        estado_cell = ws.cell(row=row, column=7, value=reg.get_estado_display())
        estado_cell.border = thin_border
        color = estado_colores.get(reg.estado, '000000')
        estado_cell.font = Font(color=color, bold=True)
        ws.cell(row=row, column=8, value=reg.get_metodo_registro_display()).border = thin_border

    for col in range(1, 9):
        ws.column_dimensions[chr(64 + col)].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="reporte_asistencias_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@login_required
def reporte_asistencias_pdf(request):
    if request.user.rol not in ['super_admin', 'admin', 'director', 'docente']:
        messages.error(request, 'Sin permisos para generar reportes.')
        return redirect('dashboard:admin_dashboard')

    from apps.asistencia.models import RegistroAsistencia

    registros = RegistroAsistencia.objects.select_related('alumno', 'curso').all()

    curso_id = request.GET.get('curso')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    if curso_id:
        registros = registros.filter(curso_id=curso_id)
    if fecha_inicio:
        registros = registros.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        registros = registros.filter(fecha__lte=fecha_fin)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'CustomTitle', parent=styles['Title'],
        fontSize=16, textColor=colors.HexColor('#1B4F72'),
        spaceAfter=6
    )
    subtitle_style = ParagraphStyle(
        'SubTitle', parent=styles['Normal'],
        fontSize=10, alignment=TA_CENTER, textColor=colors.grey,
        spaceAfter=20
    )

    elements.append(Paragraph('INSTITUTO SUPERIOR TECNOLOGICO PUBLICO RECUAY', title_style))
    elements.append(Paragraph(f'Reporte de Asistencias - {timezone.now().strftime("%d/%m/%Y %H:%M")}', subtitle_style))
    elements.append(Spacer(1, 0.3 * inch))

    data = [['N°', 'DNI', 'Alumno', 'Curso', 'Fecha', 'Hora', 'Estado']]
    for idx, reg in enumerate(registros, 1):
        data.append([
            str(idx),
            reg.alumno.dni,
            reg.alumno.get_full_name(),
            reg.curso.nombre,
            reg.fecha.strftime('%d/%m/%Y') if reg.fecha else '',
            str(reg.hora_entrada) if reg.hora_entrada else '',
            reg.get_estado_display(),
        ])

    table = Table(data, colWidths=[0.5*inch, 0.8*inch, 2*inch, 2*inch, 1*inch, 0.8*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B4F72')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F3F4')]),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)

    doc.build(elements)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_asistencias_{timezone.now().strftime("%Y%m%d")}.pdf"'
    return response


@login_required
def reporte_estadistico(request):
    from apps.asistencia.models import RegistroAsistencia
    from apps.usuarios.models import Usuario
    from apps.academica.models import Curso, Carrera

    total_alumnos = Usuario.objects.filter(rol='alumno', is_active=True).count()
    total_docentes = Usuario.objects.filter(rol='docente', is_active=True).count()
    total_cursos = Curso.objects.filter(activo=True).count()
    total_carreras = Carrera.objects.filter(activa=True).count()
    hoy = timezone.now().date()

    asistencias_hoy = RegistroAsistencia.objects.filter(fecha=hoy).count()
    presentes_hoy = RegistroAsistencia.objects.filter(fecha=hoy, estado='P').count()
    tardanzas_hoy = RegistroAsistencia.objects.filter(fecha=hoy, estado='T').count()
    faltas_hoy = RegistroAsistencia.objects.filter(fecha=hoy, estado='F').count()

    porcentaje_asistencia_hoy = round((presentes_hoy / asistencias_hoy * 100) if asistencias_hoy > 0 else 0, 1)

    from django.db.models import Count, Q
    asistencias_por_estado = RegistroAsistencia.objects.values('estado').annotate(
        total=Count('id')
    ).order_by('estado')

    cursos_con_mas_faltas = RegistroAsistencia.objects.filter(
        estado='F'
    ).values('curso__nombre').annotate(
        total=Count('id')
    ).order_by('-total')[:5]

    context = {
        'total_alumnos': total_alumnos,
        'total_docentes': total_docentes,
        'total_cursos': total_cursos,
        'total_carreras': total_carreras,
        'asistencias_hoy': asistencias_hoy,
        'presentes_hoy': presentes_hoy,
        'tardanzas_hoy': tardanzas_hoy,
        'faltas_hoy': faltas_hoy,
        'porcentaje_asistencia_hoy': porcentaje_asistencia_hoy,
        'asistencias_por_estado': list(asistencias_por_estado),
        'cursos_con_mas_faltas': list(cursos_con_mas_faltas),
    }

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        from django.http import JsonResponse
        return JsonResponse(context)

    return render(request, 'reportes/reporte_estadistico.html', context)


@login_required
def generar_qr_alumno(request, pk):
    from apps.usuarios.models import Usuario
    from django.conf import settings
    import os

    if request.user.rol not in ['super_admin', 'admin', 'docente'] and request.user.pk != pk:
        messages.error(request, 'Sin permisos.')
        return redirect('dashboard:admin_dashboard')

    usuario = Usuario.objects.get(pk=pk)

    qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_H, box_size=10, border=2)
    qr.add_data(f'TOKEN:{usuario.qr_token}|DNI:{usuario.dni}|NOMBRE:{usuario.get_full_name()}')
    qr.make(fit=True)

    qr_img = qr.make_image(fill_color='#1B4F72', back_color='white')
    qr_path = os.path.join(settings.MEDIA_ROOT, 'qr_codes', f'qr_{usuario.dni}.png')
    os.makedirs(os.path.dirname(qr_path), exist_ok=True)
    qr_img.save(qr_path)

    usuario.qr_foto.name = f'certificados_qr/qr_{usuario.dni}.png'
    usuario.save()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    header_style = ParagraphStyle('Header', parent=styles['Title'], fontSize=14, textColor=colors.HexColor('#1B4F72'), spaceAfter=6, alignment=TA_CENTER)
    info_style = ParagraphStyle('Info', parent=styles['Normal'], fontSize=11, spaceAfter=4, alignment=TA_CENTER)
    label_style = ParagraphStyle('Label', parent=styles['Normal'], fontSize=9, textColor=colors.grey, alignment=TA_CENTER)

    elements.append(Paragraph('INSTITUTO SUPERIOR TECNOLOGICO PUBLICO RECUAY', header_style))
    elements.append(Spacer(1, 0.2*inch))
    elements.append(Paragraph('CREDENCIAL DE IDENTIFICACION - CODIGO QR', info_style))
    elements.append(Spacer(1, 0.3*inch))

    qr_image = Image(qr_path, width=2.5*inch, height=2.5*inch)
    elements.append(qr_image)
    elements.append(Spacer(1, 0.2*inch))

    elements.append(Paragraph(f'<b>Nombre:</b> {usuario.get_full_name()}', info_style))
    elements.append(Paragraph(f'<b>DNI:</b> {usuario.dni}', info_style))
    elements.append(Paragraph(f'<b>Rol:</b> {usuario.get_rol_display()}', info_style))
    if usuario.codigo_estudiante:
        elements.append(Paragraph(f'<b>Codigo:</b> {usuario.codigo_estudiante}', info_style))
    elements.append(Spacer(1, 0.3*inch))
    elements.append(Paragraph('Escanea este codigo QR para registrar tu asistencia', label_style))

    doc.build(elements)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="qr_{usuario.dni}_{usuario.get_full_name().replace(" ", "_")}.pdf"'
    return response


@login_required
def generar_qr_masivo(request):
    if request.user.rol not in ['super_admin', 'admin', 'director']:
        messages.error(request, 'Sin permisos para generar QR masivo.')
        return redirect('dashboard:admin_dashboard')

    from apps.usuarios.models import Usuario
    from django.conf import settings
    import os

    alumnos = Usuario.objects.filter(rol='alumno', is_active=True)
    curso_id = request.GET.get('curso')

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    header_style = ParagraphStyle('Header', parent=styles['Title'], fontSize=14, textColor=colors.HexColor('#1B4F72'), spaceAfter=6, alignment=TA_CENTER)
    info_style = ParagraphStyle('Info', parent=styles['Normal'], fontSize=9, spaceAfter=2, alignment=TA_CENTER)

    elements.append(Paragraph('INSTITUTO SUPERIOR TECNOLOGICO PUBLICO RECUAY', header_style))
    elements.append(Paragraph('QRs DE IDENTIFICACION - ALUMNOS', info_style))
    elements.append(Spacer(1, 0.2*inch))

    for i, alumno in enumerate(alumnos):
        qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_H, box_size=8, border=2)
        qr.add_data(f'TOKEN:{alumno.qr_token}|DNI:{alumno.dni}|NOMBRE:{alumno.get_full_name()}')
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color='#1B4F72', back_color='white')
        qr_path = os.path.join(settings.MEDIA_ROOT, 'qr_codes', f'qr_{alumno.dni}.png')
        os.makedirs(os.path.dirname(qr_path), exist_ok=True)
        qr_img.save(qr_path)

        if i > 0 and i % 3 == 0:
            elements.append(Spacer(1, 0.1*inch))

        qr_image = Image(qr_path, width=1.8*inch, height=1.8*inch)
        elements.append(qr_image)
        elements.append(Paragraph(f'<b>{alumno.get_full_name()}</b>', info_style))
        elements.append(Paragraph(f'DNI: {alumno.dni}', info_style))
        elements.append(Spacer(1, 0.15*inch))

        if i % 6 == 5:
            from reportlab.platypus import PageBreak
            elements.append(PageBreak())

    doc.build(elements)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="qr_masivo_alumnos_{timezone.now().strftime("%Y%m%d")}.pdf"'
    return response
